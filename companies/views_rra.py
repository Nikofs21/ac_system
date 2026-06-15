# -*- coding: utf-8 -*-
"""
Exportacion Excel formato RRA y gestion de configuracion semanal y valores de cargo.
Solo accesible para prestadores.
"""
import io
import pytz
from datetime import date, timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.db import transaction

from companies.models import Site, SiteWeekConfig, SiteCargoValor


def require_provider(view_func):
    from functools import wraps
    @wraps(view_func)
    @login_required
    def wrapper(request, *args, **kwargs):
        if request.user.actor_type != 'PROVIDER':
            return redirect('access_denied')
        return view_func(request, *args, **kwargs)
    return wrapper


# ─────────────────────────────────────────────────────────────────────────────
# PANEL DE CONFIGURACION RRA
# ─────────────────────────────────────────────────────────────────────────────

@require_provider
def rra_config(request, site_id):
    site = get_object_or_404(Site, id=site_id)

    try:
        week_config = site.week_config
    except SiteWeekConfig.DoesNotExist:
        week_config = None

    # Cargos disponibles: JobTitle de la empresa + de la obra
    from resources.models import JobTitle
    job_titles = JobTitle.objects.filter(
        company=site.company,
        is_active=True,
    ).order_by('name')

    cargo_valores = SiteCargoValor.objects.filter(
        site=site, is_active=True
    ).order_by('cargo')

    # Set de cargos ya configurados para marcarlos en el dropdown
    cargos_configurados = {cv.cargo for cv in cargo_valores}

    return render(request, 'companies/rra_config.html', {
        'site':                 site,
        'week_config':          week_config,
        'cargo_valores':        cargo_valores,
        'job_titles':           job_titles,
        'cargos_configurados':  cargos_configurados,
        'page_title':           f'Configuración RRA — {site.name}',
    })


@require_provider
def rra_week_config_save(request, site_id):
    site = get_object_or_404(Site, id=site_id)

    if request.method == 'POST':
        base_monday_str = request.POST.get('base_monday', '').strip()
        base_week_str   = request.POST.get('base_week', '').strip()
        prefix          = request.POST.get('prefix', 'sem ').strip()

        try:
            base_monday = date.fromisoformat(base_monday_str)
            base_week   = int(base_week_str)
            # Ajustar al lunes si no lo es
            if base_monday.weekday() != 0:
                base_monday = base_monday - timedelta(days=base_monday.weekday())

            SiteWeekConfig.objects.update_or_create(
                site=site,
                defaults={
                    'base_monday': base_monday,
                    'base_week':   base_week,
                    'prefix':      prefix or 'sem ',
                }
            )
            messages.success(request, 'Configuración de semanas guardada.')
        except (ValueError, TypeError) as e:
            messages.error(request, f'Error en los datos: {e}')

    return redirect(f'/prestador/rra/{site_id}/')


@require_provider
def rra_cargo_valor_save(request, site_id):
    site = get_object_or_404(Site, id=site_id)

    if request.method == 'POST':
        # Puede venir del dropdown o del campo libre
        cargo_select = request.POST.get('cargo_select', '').strip().upper()
        cargo_libre  = request.POST.get('cargo_libre', '').strip().upper()
        cargo        = cargo_libre if cargo_libre else cargo_select
        valor_hh     = request.POST.get('valor_hh', '').strip()

        if not cargo or not valor_hh:
            messages.error(request, 'Cargo y valor son obligatorios.')
            return redirect(f'/prestador/rra/{site_id}/')

        try:
            valor = float(valor_hh)
            SiteCargoValor.objects.update_or_create(
                site=site,
                cargo=cargo,
                defaults={'valor_hh': valor, 'is_active': True}
            )
            messages.success(request, f'Valor para "{cargo}" guardado: ${valor:,.0f}/HH')
        except ValueError:
            messages.error(request, 'El valor debe ser un número.')

    return redirect(f'/prestador/rra/{site_id}/')


@require_provider
@require_POST
def rra_cargo_valor_delete(request, valor_id):
    cv = get_object_or_404(SiteCargoValor, id=valor_id)
    site_id = cv.site.id
    cv.delete()
    return JsonResponse({'status': 'ok'})


# ─────────────────────────────────────────────────────────────────────────────
# EXPORTACION EXCEL RRA
# ─────────────────────────────────────────────────────────────────────────────

@require_provider
def rra_export(request, site_id):
    """
    Genera el Excel en formato RRA listo para correr el script.

    Hoja Sesiones:
      - Todas las sesiones cerradas con HH brutas (sin regla de colacion)
      - El script RRA aplica la regla de colacion internamente usando
        inicio partida / termino partida para detectar pausas
      - semana calculada automaticamente con SiteWeekConfig
      - costo HH calculado con SiteCargoValor

    Hoja Partidas:
      - Catalogo completo de partidas con presupuesto
      - Columnas Sem XX desde sem_base hasta sem_actual + 4 futuras
        para que haya espacio de carga manual de avance fisico
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    site = get_object_or_404(Site, id=site_id)

    try:
        week_config = site.week_config
    except SiteWeekConfig.DoesNotExist:
        messages.error(request, 'Esta obra no tiene configuración de semanas ISA.')
        return redirect(f'/prestador/rra/{site_id}/')

    site_tz   = pytz.timezone(site.timezone or 'America/Santiago')

    # Mapa cargo → valor_hh (normalizado a mayúsculas)
    cargo_map = {
        cv.cargo.upper(): float(cv.valor_hh)
        for cv in SiteCargoValor.objects.filter(site=site, is_active=True)
    }

    wb = openpyxl.Workbook()

    # Estilos
    header_fill = PatternFill(start_color="0025EC", end_color="0025EC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_aln  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin        = Side(style='thin', color="DDDDDD")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Hoja Sesiones ─────────────────────────────────────────────────────────
    ws_ses = wb.active
    ws_ses.title = 'Sesiones'

    ses_headers = [
        'nombre', 'cargo', 'fecha', 'semana',
        'sub etapa', 'partida', 'etapa', 'supervisor',
        'inicio partida', 'término partida',
        'HH', 'costo HH', 'tipo',
    ]

    for col_idx, h in enumerate(ses_headers, start=1):
        cell = ws_ses.cell(row=1, column=col_idx, value=h)
        cell.fill   = header_fill
        cell.font   = header_font
        cell.alignment = header_aln
        cell.border = border

    # Leer sesiones cerradas
    from work.models import WorkSession, StageTask
    sessions = WorkSession.objects.filter(
        site=site,
        status__in=['CLOSED', 'AUTO_CLOSED'],
        ended_at__isnull=False,
    ).select_related(
        'resource', 'resource__job_title',
        'stage', 'task',
        'started_by',
        'responsible_supervisor',
    ).order_by('started_at')

    # Cache de StageTask para no hacer query por sesion
    stage_task_cache = {}
    for st in StageTask.objects.filter(site=site).select_related('stage', 'task'):
        key = (st.stage_id, st.task_id)
        stage_task_cache[key] = st

    for row_idx, s in enumerate(sessions, start=2):
        inicio_local = s.started_at.astimezone(site_tz)
        fin_local    = s.ended_at.astimezone(site_tz)
        fecha_local  = inicio_local.date()

        nombre = s.resource.display_name
        cargo  = s.resource.job_title.name.upper() if s.resource.job_title else ''

        # Semana ISA
        semana = week_config.get_week_label_for_date(fecha_local)

        # Subetapa y tipo desde StageTask
        st = stage_task_cache.get((s.stage_id, s.task_id))
        subetapa = st.subetapa if st and st.subetapa else ''
        tipo     = st.tipo if st else 'casa'

        partida = s.task_name_snapshot
        etapa   = s.stage_name_snapshot

        # Supervisor
        if s.responsible_supervisor:
            supervisor = s.responsible_supervisor.get_full_name() or s.responsible_supervisor.email
        elif s.started_by:
            supervisor = s.started_by.get_full_name() or s.started_by.email
        else:
            supervisor = ''

        # HH brutas — el script aplica la regla de colacion internamente
        hh = round((s.ended_at - s.started_at).total_seconds() / 3600, 4)

        # Costo HH
        valor    = cargo_map.get(cargo, None)
        costo_hh = round(hh * valor, 2) if valor is not None else ''

        row = [
            nombre,
            cargo,
            fecha_local,
            semana,
            subetapa,
            partida,
            etapa,
            supervisor,
            inicio_local.replace(tzinfo=None),
            fin_local.replace(tzinfo=None),
            hh,
            costo_hh,
            tipo,
        ]

        for col_idx, val in enumerate(row, start=1):
            cell = ws_ses.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9FF", end_color="F9F9FF", fill_type="solid")

    # Formatos de fecha
    for row in ws_ses.iter_rows(min_row=2, max_row=ws_ses.max_row):
        row[2].number_format  = 'DD/MM/YYYY'
        row[8].number_format  = 'DD/MM/YYYY HH:MM'
        row[9].number_format  = 'DD/MM/YYYY HH:MM'
        row[10].number_format = '0.0000'
        if row[11].value != '':
            row[11].number_format = '#,##0'

    ses_widths = [28, 22, 12, 10, 38, 48, 32, 24, 18, 18, 9, 12, 12]
    for i, w in enumerate(ses_widths, start=1):
        ws_ses.column_dimensions[get_column_letter(i)].width = w

    ws_ses.row_dimensions[1].height = 32
    ws_ses.freeze_panes = 'A2'

    # ── Hoja Partidas ─────────────────────────────────────────────────────────
    ws_par = wb.create_sheet('Partidas')

    # Calcular rango de semanas
    # Desde base_week hasta semana actual + 4 futuras (espacio para avance)
    today        = date.today()
    current_week = week_config.get_week_for_date(today)
    sem_inicio   = week_config.base_week
    sem_fin      = current_week + 4

    sem_cols = [f'Sem {w}' for w in range(sem_inicio, sem_fin + 1)]

    par_headers = [
        'etapa', 'subetapa', 'partida', 'partida_cod',
        'estado_partida', 'tipo',
        'cantidad presupuesto', 'unidad medida',
        'presupuesto para mano obra directa',
        'presupuesto total',
    ] + sem_cols

    for col_idx, h in enumerate(par_headers, start=1):
        cell = ws_par.cell(row=1, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_aln
        cell.border    = border

    stage_tasks = StageTask.objects.filter(
        site=site
    ).select_related('stage', 'task').order_by(
        'stage__name', 'display_order', 'task__name'
    )

    for row_idx, st in enumerate(stage_tasks, start=2):
        row_data = [
            st.stage.name,
            st.subetapa or '',
            st.task.name,
            st.partida_cod or st.task.name,
            st.estado_partida,
            st.tipo,
            float(st.cantidad_presupuesto)  if st.cantidad_presupuesto  else '',
            st.unidad_medida or '',
            float(st.presupuesto_mo)        if st.presupuesto_mo        else '',
            float(st.presupuesto_total)     if st.presupuesto_total     else '',
        ] + ['' for _ in sem_cols]  # columnas Sem XX vacías para llenar a mano

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws_par.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(start_color="F9F9FF", end_color="F9F9FF", fill_type="solid")

    par_base_widths = [32, 38, 48, 52, 13, 14, 18, 12, 30, 18]
    for i, w in enumerate(par_base_widths, start=1):
        ws_par.column_dimensions[get_column_letter(i)].width = w
    for i in range(len(par_base_widths) + 1, len(par_headers) + 1):
        ws_par.column_dimensions[get_column_letter(i)].width = 10

    ws_par.row_dimensions[1].height = 32
    ws_par.freeze_panes = 'A2'

    # ── Descargar ──────────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'RRA_{site.code}_{site.name.replace(" ", "_")}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
