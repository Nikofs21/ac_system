# -*- coding: utf-8 -*-
"""
Vista de revision de partidas.
Agregar estas funciones al final de work/views.py
"""
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.utils import timezone
from django.db import transaction
from datetime import date, timedelta

from resources.models import Resource, ResourceSiteAssignment
from tracking.models import NoOnSiteEvent
from .models import Stage, TaskCatalog, StageTask, WorkSession, WorkSessionChangeLog
from core.permissions import user_has_permission, get_user_context_permissions
from core.utils import get_active_site


def require_active_site(view_func):
    from functools import wraps
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if not get_active_site(request):
            return redirect('select_site')
        return view_func(request, *args, **kwargs)
    return login_required(wrapper)


@require_active_site
def session_review(request):
    """
    Vista principal de revision de partidas.
    Muestra sesiones del dia seleccionado y activos sin sesion.
    """
    site = get_active_site(request)

    if not user_has_permission(request.user, 'sessions_review.view', site):
        return redirect('access_denied')

    # Fecha seleccionada (default: hoy)
    date_str = request.GET.get('fecha', '')
    try:
        selected_date = date.fromisoformat(date_str)
    except ValueError:
        selected_date = timezone.localdate()

    # Busqueda por nombre o RUT
    query = request.GET.get('q', '').strip()

    today = timezone.localdate()
    is_today = (selected_date == today)

    # Recursos asignados a la obra
    assignments = ResourceSiteAssignment.objects.filter(
        site=site,
        status='ACTIVE',
        resource__status='ACTIVE',
    ).select_related(
        'resource',
        'resource__job_title',
    )

    # Filtrar por busqueda si hay query
    if query:
        from django.db.models import Q
        assignments = assignments.filter(
            Q(resource__display_name__icontains=query) |
            Q(resource__normalized_name__icontains=query.lower()) |
            Q(resource__person_rut__icontains=query.replace('.', '').replace('-', ''))
        )

    resource_ids = list(assignments.values_list('resource_id', flat=True))

    # Sesiones del dia seleccionado
    sessions_qs = WorkSession.objects.filter(
        site=site,
        resource_id__in=resource_ids,
        started_at__date=selected_date,
    ).select_related(
        'resource', 'resource__job_title',
        'task', 'stage',
        'started_by', 'ended_by',
        'responsible_supervisor',
    ).order_by('resource__display_name', 'started_at')

    # Agrupar sesiones por recurso
    sessions_by_resource = {}
    for session in sessions_qs:
        rid = session.resource_id
        if rid not in sessions_by_resource:
            sessions_by_resource[rid] = {
                'resource': session.resource,
                'sessions': [],
            }
        sessions_by_resource[rid]['sessions'].append(session)

    # Marcas No en obra del dia
    nos_map = {
        e.resource_id: e
        for e in NoOnSiteEvent.objects.filter(
            site=site,
            event_date=selected_date,
            status='ACTIVE',
        ).select_related('marked_by')
    }

    # Construir lista de trabajadores con sesiones
    workers_with_sessions = []
    workers_without_sessions = []

    for assignment in assignments.order_by('resource__display_name'):
        resource = assignment.resource
        rid = resource.id
        sessions_data = sessions_by_resource.get(rid)
        nos_event = nos_map.get(rid)

        # Estado general del trabajador para el dia
        if sessions_data:
            has_open = any(s.status == 'OPEN' for s in sessions_data['sessions'])
            estado = 'open' if has_open else 'closed'
            workers_with_sessions.append({
                'resource':  resource,
                'sessions':  sessions_data['sessions'],
                'estado':    estado,
                'nos_event': nos_event,
            })
        else:
            workers_without_sessions.append({
                'resource':  resource,
                'nos_event': nos_event,
            })

    total    = len(resource_ids)
    con_sesion = len(workers_with_sessions)
    sin_sesion = len(workers_without_sessions)

    perms_ctx = get_user_context_permissions(request.user, site)

    return render(request, 'work/session_review.html', {
        'workers_with_sessions':    workers_with_sessions,
        'workers_without_sessions': workers_without_sessions,
        'selected_date':   selected_date,
        'today':           today,
        'is_today':        is_today,
        'query':           query,
        'site':            site,
        'page_title':      'Revision de partidas',
        'total':           total,
        'con_sesion':      con_sesion,
        'sin_sesion':      sin_sesion,
        'perms_ctx':       perms_ctx,
        'can_edit':        perms_ctx.get('can_edit_today', False) and is_today,
    })


@require_active_site
def session_edit(request, session_id):
    """
    Editar etapa y partida de una sesion del dia actual.
    Solo disponible para sesiones de hoy.
    Solo cambia stage y task, no horarios ni personas.
    """
    site = get_active_site(request)

    if not user_has_permission(request.user, 'sessions_review.edit_today', site):
        return JsonResponse({'error': 'Sin permiso para editar sesiones.'}, status=403)

    session = get_object_or_404(
        WorkSession,
        id=session_id,
        site=site,
    )

    today = timezone.localdate()

    # Solo sesiones del dia actual
    if session.started_at.date() != today:
        return JsonResponse({'error': 'Solo se pueden editar sesiones del dia actual.'}, status=400)

    if request.method == 'GET':
        # Retornar datos actuales + opciones disponibles
        stage_tasks = StageTask.objects.filter(
            site=site,
            is_active=True,
            stage__is_active=True,
            task__status='ACTIVE',
            stage__stage_type='NORMAL',
        ).select_related('stage', 'task').order_by('stage__name', 'task__name')

        etapas = {}
        for st in stage_tasks:
            sid = st.stage.id
            if sid not in etapas:
                etapas[sid] = {
                    'id':   st.stage.id,
                    'name': st.stage.name,
                    'tasks': [],
                }
            etapas[sid]['tasks'].append({
                'id':         st.task.id,
                'name':       st.task.name,
                'risk_level': st.task.risk_level,
            })

        return JsonResponse({
            'session_id':   session.id,
            'current_stage_id': session.stage_id,
            'current_task_id':  session.task_id,
            'stage_name':   session.stage_name_snapshot,
            'task_name':    session.task_name_snapshot,
            'worker_name':  session.resource.display_name,
            'started_at':   session.started_at.strftime('%H:%M'),
            'status':       session.status,
            'etapas':       list(etapas.values()),
        })

    if request.method == 'POST':
        new_stage_id = request.POST.get('stage_id')
        new_task_id  = request.POST.get('task_id')
        reason       = request.POST.get('reason', '').strip()

        if not new_stage_id or not new_task_id:
            return JsonResponse({'error': 'Etapa y partida son obligatorias.'}, status=400)

        new_stage = get_object_or_404(Stage, id=new_stage_id, site=site)
        new_task  = get_object_or_404(TaskCatalog, id=new_task_id, company=site.company)

        # Verificar que la combinacion etapa/partida existe en la obra
        valid = StageTask.objects.filter(
            site=site,
            stage=new_stage,
            task=new_task,
            is_active=True,
        ).exists()

        if not valid:
            return JsonResponse({'error': 'Combinacion de etapa y partida no valida para esta obra.'}, status=400)

        with transaction.atomic():
            # Guardar estado anterior para auditoria
            before = {
                'stage_id':            session.stage_id,
                'task_id':             session.task_id,
                'stage_name_snapshot': session.stage_name_snapshot,
                'task_name_snapshot':  session.task_name_snapshot,
                'task_code_snapshot':  session.task_code_snapshot,
                'risk_level_snapshot': session.risk_level_snapshot,
            }

            # Aplicar cambios
            session.stage               = new_stage
            session.task                = new_task
            session.stage_name_snapshot = new_stage.name
            session.task_name_snapshot  = new_task.name
            session.task_code_snapshot  = new_task.code
            session.risk_level_snapshot = new_task.risk_level
            session.save()

            # Registrar en bitacora
            WorkSessionChangeLog.objects.create(
                session=session,
                changed_by=request.user,
                change_type='DAY_CORRECTION',
                before_json=before,
                after_json={
                    'stage_id':            new_stage.id,
                    'task_id':             new_task.id,
                    'stage_name_snapshot': new_stage.name,
                    'task_name_snapshot':  new_task.name,
                    'task_code_snapshot':  new_task.code,
                    'risk_level_snapshot': new_task.risk_level,
                },
                reason=reason or 'Correccion del dia',
            )

        return JsonResponse({
            'status':          'ok',
            'new_stage_name':  new_stage.name,
            'new_task_name':   new_task.name,
            'new_risk_level':  new_task.risk_level,
        })

    return JsonResponse({'error': 'Metodo no permitido.'}, status=405)

@require_active_site
def session_review_export(request):
    """
    Exporta sesiones en rango de fechas al formato Excel de revision de partidas.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import pytz

    site = get_active_site(request)

    if not user_has_permission(request.user, 'sessions_review.view', site):
        return redirect('access_denied')

    # Rango de fechas
    date_from_str = request.GET.get('fecha_desde', '')
    date_to_str   = request.GET.get('fecha_hasta', '')

    try:
        date_from = date.fromisoformat(date_from_str)
    except ValueError:
        date_from = timezone.localdate() - timedelta(days=7)

    try:
        date_to = date.fromisoformat(date_to_str)
    except ValueError:
        date_to = timezone.localdate()

    # No permitir rangos mayores a 90 días
    if (date_to - date_from).days > 90:
        date_from = date_to - timedelta(days=90)

    site_tz = pytz.timezone(site.timezone or 'America/Santiago')

    # Obtener sesiones del rango
    from work.models import StageTask
    sessions = WorkSession.objects.filter(
        site=site,
        started_at__date__gte=date_from,
        started_at__date__lte=date_to,
        status__in=['CLOSED', 'AUTO_CLOSED'],
    ).select_related(
        'resource', 'resource__job_title',
        'stage', 'task',
        'responsible_supervisor',
    ).order_by('started_at', 'resource__display_name')

    # Cache StageTask para subetapa
    st_cache = {}
    for st in StageTask.objects.filter(site=site).select_related('stage', 'task'):
        st_cache[(st.stage_id, st.task_id)] = st

    # Construir Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Sesiones'

    # Estilos
    header_fill = PatternFill(start_color="0025EC", end_color="0025EC", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    header_aln  = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin        = Side(style='thin', color="DDDDDD")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    headers = [
        'Fecha', 'Nombre', 'Cargo', 'Etapa',
        'Subetapa', 'Partida',
        'Inicio', 'Término', 'Duración (min)',
        'Supervisor', 'Estado', 'Subetapa — Partida',
    ]

    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = header_aln
        cell.border    = border

    for row_idx, s in enumerate(sessions, start=2):
        inicio_local = s.started_at.astimezone(site_tz)
        fin_local    = s.ended_at.astimezone(site_tz) if s.ended_at else None
        fecha_local  = inicio_local.date()

        nombre = s.resource.display_name
        cargo  = s.resource.job_title.name if s.resource.job_title else ''
        etapa  = s.stage_name_snapshot
        partida = s.task_name_snapshot

        st = st_cache.get((s.stage_id, s.task_id))
        subetapa = st.subetapa if st and st.subetapa else ''

        supervisor = ''
        if s.responsible_supervisor:
            supervisor = s.responsible_supervisor.get_full_name() or s.responsible_supervisor.email

        dur_min = s.duration_minutes or (
            int((s.ended_at - s.started_at).total_seconds() / 60)
            if s.ended_at else ''
        )

        estado = 'Cerrada' if s.status == 'CLOSED' else 'Cierre automático'

        concat = f'{subetapa} — {partida}' if subetapa else partida

        row = [
            fecha_local,
            nombre,
            cargo,
            etapa,
            subetapa,
            partida,
            inicio_local.replace(tzinfo=None),
            fin_local.replace(tzinfo=None) if fin_local else '',
            dur_min,
            supervisor,
            estado,
            concat,
        ]

        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if row_idx % 2 == 0:
                cell.fill = PatternFill(
                    start_color="F9F9FF", end_color="F9F9FF", fill_type="solid"
                )

    # Formatos de fecha y hora
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        row[0].number_format = 'DD/MM/YYYY'
        row[6].number_format = 'DD/MM/YYYY HH:MM'
        row[7].number_format = 'DD/MM/YYYY HH:MM'

    # Anchos de columna
    widths = [12, 28, 20, 30, 35, 45, 18, 18, 14, 22, 16, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 30
    ws.freeze_panes = 'A2'

    # Descargar
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from django.http import HttpResponse
    filename = f'Revision_{site.code}_{date_from}_{date_to}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
